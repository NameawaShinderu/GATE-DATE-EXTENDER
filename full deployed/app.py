import openpyxl
from datetime import datetime, timedelta
from flask import Flask, render_template, request

app = Flask(__name__)

# Function to increase the dates by a given value
def increase_dates(date_list, increment):
    updated_dates = []
    for date_range in date_list:
        if date_range == "Don't do":
            updated_dates.append(date_range)
            continue
        if '-' in date_range:
            start_date, end_date = date_range.split(' - ')
            start_date = datetime.strptime(start_date, '%B %d')
            end_date = datetime.strptime(end_date, '%B %d')
            date_increment = timedelta(days=increment)
            updated_start_date = start_date + date_increment
            updated_end_date = end_date + date_increment
            updated_date_range = f"{updated_start_date.strftime('%B %d')} - {updated_end_date.strftime('%B %d')}"
        else:
            date = datetime.strptime(date_range, '%B %d')
            date_increment = timedelta(days=increment)
            updated_date_range = (date + date_increment).strftime('%B %d')
        updated_dates.append(updated_date_range)
    return updated_dates

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        user_date_input = request.form['user_date']
        try:
            # Convert the user input to a datetime object
            user_date = datetime.strptime(user_date_input, '%Y-%m-%d')
            
            # Load the Excel file
            workbook = openpyxl.load_workbook('example.xlsx')
            
            # Select the desired sheet (you may need to modify 'Sheet1' to match your sheet name)
            sheet = workbook['Sheet1']
            
            # Define the dataset of dates
            dates = [
                'July 8 - July 10',
    'July 11 - July 13',
    'July 14',
    'July 15',
    'July 16',
    'July 17 - July 19',
    'July 20',
    'July 21 - July 23',
    'July 24',
    'July 25',
    'July 25',
    'July 26 - July 29',
    'July 30',
    'July 31',
    'August 1 - August 2',
    'August 3',
    'August 4 - August 5',
    'August 6',
    'August 7 - August 8',
    'August 9',
    'August 10',
    'August 10',
    'August 11',
    'August 12',
    'August 13',
    'August 14',
    'August 15',
    'August 16',
    'August 17',
    'August 18',
    'August 19 - August 20',
    'August 21',
    'August 21',
    'August 22 - August 23',
    'August 24',
    'August 25 - August 27',
    'August 28',
    'August 29 - August 30',
    'August 31',
    'September 1 - September 2',
    'September 3',
    'September 4 - September 5',
    'September 6',
    'September 7 - September 9',
    'September 10',
    'September 11',
    'September 12',
    'September 13',
    'September 13',
    'September 14 - September 15',
    'September 16 - September 17',
    'September 18 - September 19',
    'September 20 - September 21',
    'September 22 - September 23',
    'September 24 - September 25',
    'September 26 - September 27',
    'September 28 - September 29',
    'September 30 - October 1',
    'October 1 - October 4',
    'October 5 - October 7',
    'October 8',
    "Don't do",
    'October 9 - October 11',
    'October 12',
    'October 13 - October 16',
    'October 17',
    "Don't do",
    'October 18 - October 19',
    'October 20 - October 21',
    'October 22'
                # Your list of dates here
            ]
            
            # Calculate the difference in days
            july_8 = datetime.strptime('July 8', '%B %d')
            difference = (user_date - july_8).days
            
            # Example usage with the difference as the increment value
            updated_dates = increase_dates(dates, difference)
            
            # Iterate through the updated dates and update the "dates" column
            for i, date_range in enumerate(updated_dates, start=2):
                cell = f'C{i}'  # Assuming the "dates" column starts at column C
                sheet[cell] = date_range
            
            # Save the modified Excel file
            workbook.save('your_file_updated.xlsx')

            def read_excel(file_path):
              workbook = openpyxl.load_workbook(file_path)
              sheet = workbook.active
              data = []
              for row in sheet.iter_rows(values_only=True):
                 data.append(row)
              return data

            

            updated_data = read_excel('your_file_updated.xlsx')
            
            return render_template('result.html', excel_data=updated_data)
                    
        except ValueError:
            return render_template('error.html')
    
    return render_template('index.html')

@app.route('/today', methods=['GET'])
def today():
    try:
        # Get the current date
        current_date = datetime.now().date()
        
        # Load the Excel file
        workbook = openpyxl.load_workbook('example.xlsx')
        
        # Select the desired sheet (you may need to modify 'Sheet1' to match your sheet name)
        sheet = workbook['Sheet1']
        
        # Define the dataset of dates
        dates = [
            'July 8 - July 10',
    'July 11 - July 13',
    'July 14',
    'July 15',
    'July 16',
    'July 17 - July 19',
    'July 20',
    'July 21 - July 23',
    'July 24',
    'July 25',
    'July 25',
    'July 26 - July 29',
    'July 30',
    'July 31',
    'August 1 - August 2',
    'August 3',
    'August 4 - August 5',
    'August 6',
    'August 7 - August 8',
    'August 9',
    'August 10',
    'August 10',
    'August 11',
    'August 12',
    'August 13',
    'August 14',
    'August 15',
    'August 16',
    'August 17',
    'August 18',
    'August 19 - August 20',
    'August 21',
    'August 21',
    'August 22 - August 23',
    'August 24',
    'August 25 - August 27',
    'August 28',
    'August 29 - August 30',
    'August 31',
    'September 1 - September 2',
    'September 3',
    'September 4 - September 5',
    'September 6',
    'September 7 - September 9',
    'September 10',
    'September 11',
    'September 12',
    'September 13',
    'September 13',
    'September 14 - September 15',
    'September 16 - September 17',
    'September 18 - September 19',
    'September 20 - September 21',
    'September 22 - September 23',
    'September 24 - September 25',
    'September 26 - September 27',
    'September 28 - September 29',
    'September 30 - October 1',
    'October 1 - October 4',
    'October 5 - October 7',
    'October 8',
    "Don't do",
    'October 9 - October 11',
    'October 12',
    'October 13 - October 16',
    'October 17',
    "Don't do",
    'October 18 - October 19',
    'October 20 - October 21',
    'October 22'
            # Your list of dates here
        ]
        
        # Calculate the difference in days
        july_8 = datetime.strptime('July 8', '%B %d')
        difference = (current_date - july_8.date()).days
        
        # Example usage with the difference as the increment value
        updated_dates = increase_dates(dates, difference)
        
        # Iterate through the updated dates and update the "dates" column
        for i, date_range in enumerate(updated_dates, start=2):
            cell = f'C{i}'  # Assuming the "dates" column starts at column C
            sheet[cell] = date_range
        
        # Save the modified Excel file
        workbook.save('your_file_updated.xlsx')

        def read_excel(file_path):
              workbook = openpyxl.load_workbook(file_path)
              sheet = workbook.active
              data = []
              for row in sheet.iter_rows(values_only=True):
                 data.append(row)
              return data

            

        updated_data = read_excel('your_file_updated.xlsx')
            
        return render_template('result.html', excel_data=updated_data)
        
    
    except ValueError:
        return render_template('error.html')

if __name__ == '__main__':
    app.run(debug=True)
